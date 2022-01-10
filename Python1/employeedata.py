''' openyxl is a python libary that permits us to read and edit xl files '''
import openpyxl as xl

wb = xl.load_workbook('employeesdatabase.xlsx')

sheet = wb['Sheet1']

old_email = 'helpinghands.cm'
new_email = 'handsinhands.org'

for i in range(2, sheet.max_row+1):
    cell = sheet.cell(i, 2)
    
    if old_email in cell.value:
        updated_Email = (cell.value).replace(old_email , new_email)
        
        sheet.cell(i,2).value = updated_Email
wb.save('employeesdata.csv')